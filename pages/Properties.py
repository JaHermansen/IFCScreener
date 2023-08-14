# Quantities

import streamlit as st
from tools import ifchelper
from tools import pandashelper
from tools import graph_maker
import pandas as pd
from PIL import Image
import ifcopenshell
import ifcopenshell.util.element as Element
from io import BytesIO
import xlsxwriter
from datetime import date
from pathlib import Path
import os
from openpyxl import load_workbook


#Page icon
icon = Image.open('Images/paa1.png')

st.set_page_config(
    layout= "wide",
    page_title= "IFC Screener",
    page_icon= icon,
)


session = st.session_state
# Get the original file name
original_file_name = st.session_state["file_name"]
# Create the updated file name with the current date
updated_file_name = original_file_name.split(".")[0] + "_updated_" + str(date.today()) + ".ifc"



def get_ifc_pandas():
    data, pset_attributes = ifchelper.get_objects_data_by_class(
        session.ifc_file,
        "IfcBuildingElement"
        )
    return ifchelper.create_pandas_dataframe(data, pset_attributes)

def get_ifc_pandas_filter(class_filter):
    data, pset_attributes = ifchelper.get_objects_data_by_class(
        session.ifc_file,
        class_filter
        )
    return ifchelper.create_pandas_dataframe(data, pset_attributes)

def explore():
    # DATA
    st.write("Data:")
    st.write(session.df)  # use df from session_state

def get_df(file):
    # get extension and read file
    extension = file.name.split('.')[1]
    if extension.upper() == 'CSV':
        df = pd.read_csv(file)
    elif extension.upper() == 'XLSX':
        df = pd.read_excel(file, engine='openpyxl')
    elif extension.upper() == 'PICKLE':
        df = pd.read_pickle(file)
    #st.write(f"Debug: {df.head()}")  # debug line
    session.df = df  # store df in session_state

def download_csv():
    pandashelper.download_csv(session.file_name, session.Dataframe)
    
def download_excel():
    pandashelper.download_excel(session.file_name,session.DataFrame)

def callback_upload():
    st.session_state["file_name"] = st.session_state["uploaded_file"].name
    st.session_state["array_buffer"] = st.session_state["uploaded_file"].getvalue()
    st.session_state["is_file_uploaded"] = True

def update_properties(bim_type_codes_selected):

    owner_history = session.ifc_file.by_type("IfcOwnerHistory")[0]
    products = session.ifc_file.by_type("IfcProduct")
    procs = [i for i in products if i.is_a("IfcProduct")]

    for proc in procs:
        property_sets = ifcopenshell.util.element.get_psets(proc)
        property_values = []
        for bim_type_code in bim_type_codes_selected:
            pset_value = None
            for pset_name, properties in property_sets.items():
                for prop_name, prop_values in properties.items():
                    if bim_type_code in prop_name:
                        pset_value = prop_values
                        break
                if pset_value is not None:
                    break

            if pset_value is not None:
                property_values.append(
                    session.ifc_file.createIfcPropertySingleValue(
                        bim_type_code, bim_type_code, session.ifc_file.create_entity("IfcText", str(pset_value)), None
                    )
                )
            else:
                property_values.append(
                    session.ifc_file.createIfcPropertySingleValue(
                        bim_type_code, bim_type_code, session.ifc_file.create_entity("IfcText", "Not Available"), None
                    )
                )
        property_set = session.ifc_file.createIfcPropertySet(proc.GlobalId, owner_history, "Pset_PAABIMTypeCodes", None, property_values)
        session.ifc_file.createIfcRelDefinesByProperties(proc.GlobalId, owner_history, None, None, [proc], property_set)

    # Write the modified IFC file to the Downloads folder
    #downloads_path = Path.home() / "Downloads"
    #updated_file_path = downloads_path.joinpath(updated_file_name)
    #session.ifc_file.write(str(updated_file_path))

    # Write the modified IFC file to a temporary file on disk
    temp_file_path = "temp_updated_file.ifc"
    session.ifc_file.write(temp_file_path)

    # Read the file into a BytesIO object
    updated_file_bytes = BytesIO()
    with open(temp_file_path, 'rb') as f:
        updated_file_bytes.write(f.read())
    updated_file_bytes.seek(0)

    # Optional: Delete the temporary file if it's no longer needed
    os.remove(temp_file_path)

    return updated_file_bytes


def add_new_properties_old(new_properties_dict):
    owner_history = session.ifc_file.by_type("IfcOwnerHistory")[0]
    products = session.ifc_file.by_type("IfcProduct")
    procs = [i for i in products if i.is_a("IfcProduct")]

    for proc in procs:
        identifier = proc.GlobalId  # Get the GlobalId of the current product

        if identifier in new_properties_dict:
            element_properties = new_properties_dict[identifier]
            property_values = []
            for col, values in element_properties.items():
                if "." in col:
                    property_set_name, property_name = col.split(".", 1)
                else:
                    property_set_name = col
                    property_name = col
                property_values.extend(
                    session.ifc_file.createIfcPropertySingleValue(
                        property_name, property_name, session.ifc_file.create_entity("IfcText", str(value)), None
                    )
                    for value in values
                )

            property_set = session.ifc_file.createIfcPropertySet(
                identifier, owner_history, property_set_name, None, property_values
            )
            session.ifc_file.createIfcRelDefinesByProperties(identifier, owner_history, None, None, [proc], property_set)

    # Write the modified IFC file to the Downloads folder
    downloads_path = Path.home() / "Downloads"
    updated_file_path = downloads_path.joinpath(updated_file_name)
    session.ifc_file.write(str(updated_file_path))

def add_new_properties(new_properties_dict):
    owner_history = session.ifc_file.by_type("IfcOwnerHistory")[0]
    products = session.ifc_file.by_type("IfcProduct")
    procs = [i for i in products if i.is_a("IfcProduct")]

    for proc in procs:
        identifier = proc.GlobalId  # Get the GlobalId of the current product

        if identifier in new_properties_dict:
            element_properties = new_properties_dict[identifier]

            for col, values in element_properties.items():
                if "." in col:
                    property_set_name, property_name = col.split(".", 1)
                else:
                    property_set_name = col
                    property_name = col
                
                property_values = []
                for value in values:
                    existing_property = None
                    for prop_set in proc.IsDefinedBy:
                        if (
                            prop_set.is_a("IfcRelDefinesByProperties")
                            and prop_set.RelatingPropertyDefinition.Name == property_set_name
                        ):
                            for prop in prop_set.RelatingPropertyDefinition.HasProperties:
                                if prop.Name == property_name:
                                    existing_property = prop
                                    break
                            if existing_property:
                                break

                    if existing_property:
                        # Update the existing property's value
                        existing_property.NominalValue = session.ifc_file.create_entity("IfcText", str(value))
                    else:
                        # Create a new property
                        property_values.append(
                            session.ifc_file.createIfcPropertySingleValue(
                                property_name, property_name,
                                session.ifc_file.create_entity("IfcText", str(value)), None
                            )
                        )
                
                if property_values:
                    property_set = session.ifc_file.createIfcPropertySet(
                        identifier, owner_history, property_set_name, None, property_values
                    )
                    session.ifc_file.createIfcRelDefinesByProperties(
                        identifier, owner_history, None, None, [proc], property_set
                    )


    # Write the modified IFC file to the Downloads folder
    #downloads_path = Path.home() / "Downloads"
    #updated_file_path = downloads_path.joinpath(updated_file_name)
    #session.ifc_file.write(str(updated_file_path))

    # Write the modified IFC file to a temporary file on disk
    tempo_file_path = "tempo_updated_file.ifc"
    session.ifc_file.write(tempo_file_path)

    # Read the file into a BytesIO object
    updated_file = BytesIO()
    with open(tempo_file_path, 'rb') as f:
        updated_file.write(f.read())
    updated_file.seek(0)

    # Optional: Delete the temporary file if it's no longer needed
    os.remove(tempo_file_path)

    return updated_file

def find_sheet_with_class(sheet_names, ifc_building_element, file):
    for sheet_name in sheet_names:
        df_sheet = pd.read_excel(file, sheet_name)
        global_ids = df_sheet["GlobalId"].tolist()

        if ifc_building_element.GlobalId in global_ids:
            return sheet_name

    return None
    
def read_data_from_excel(df, sheet_name, global_id_column):
    # Read the sheet with the specified name
    sheet = pd.read_excel(df, sheet_name)

    # Read the IfcBuildingElement objects from the sheet
    ifc_building_elements = session.ifc_file.by_type("IfcBuildingElement")

    # Create a dictionary to store the associated rows of data
    data_dict = {}

    # Iterate over each object in IfcBuildingElement class
    for element in ifc_building_elements:
        global_id = element.GlobalId

        # Find the row in the sheet with the matching GlobalId
        row = sheet[sheet[global_id_column] == global_id]

        # If a matching row is found, store it in the data dictionary
        if not row.empty:
            data_dict[global_id] = row.iloc[0]

    return data_dict

def find_row_with_global_id(df_sheet, global_id):
    row = df_sheet[df_sheet["GlobalId"] == global_id]
    return row.iloc[0] if not row.empty else None

def compare_datasets(df1, df2, identifier_column):
    df1_columns = set(df1.columns)
    df2_columns = set(df2.columns)
    new_columns = df1_columns - df2_columns - {identifier_column}
    return new_columns


def compare_specific_values(df1, df2, specific_columns, identifier_column):
    differing_values_dict = {}  # Dictionary to store differing values by GlobalId

    common_columns = set(specific_columns).intersection(df1.columns).intersection(df2.columns)

    for index, row1 in df1.iterrows():
        global_id = row1[identifier_column]
        row2 = df2[df2[identifier_column] == global_id]

        if not row2.empty:
            differing_props = {}
            for col in common_columns:
                col_name = col.split(".")[-1]
                value1 = row1[col]
                value2 = row2[col].iloc[0]

                if value1 != value2:
                    differing_props[col] = [value2]

            if differing_props:
                differing_values_dict[global_id] = differing_props

    return differing_values_dict




def execute():
    st.markdown("<h1 style='color: #006095;'>Model Properties</h1>", unsafe_allow_html=True)
    if st.session_state.get("ifc_file") is None:
        st.warning("No file provided. Please upload a file.")
    else:
        
        tab1, tab2, tab3, tab4 = st.tabs(["Properties Overview", "Quantities Review", "BIMTypeCodes", "Add External Data"])
        
        with tab1:

            
            st.header("Properties Overview")
            st.write("Overall dataframe")
            session["DataFrame"] = get_ifc_pandas()
            st.write(session["DataFrame"])

            ### Pre-create dataframe for export
            output = BytesIO()
            dataframe = session["DataFrame"]

            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            CLASS = "Class"

            for object_class in dataframe[CLASS].unique():
                df_class = dataframe[dataframe[CLASS] == object_class].dropna(axis=1, how="all")

                worksheet = workbook.add_worksheet(object_class)  # create worksheet with name 'object_class'

                for r_idx, row in enumerate(df_class.values):
                    for c_idx, value in enumerate(row):
                        worksheet.write(r_idx + 1, c_idx, value)  # +1 as headers are already written

                for idx, col in enumerate(df_class):  # loop through all columns
                    series = df_class[col]
                    if "PAA" in col:
                        format = workbook.add_format({'bg_color': '#ADD8E6'})
                        # Create the cell range for the column
                        cell_range = xlsxwriter.utility.xl_range(1, idx, len(series), idx)
                        worksheet.conditional_format(cell_range, {'type': 'no_blanks', 'format': format})

            # Write the column headers
            for c_idx, col_name in enumerate(df_class.columns):
                worksheet.write(0, c_idx, col_name)

            workbook.close()

            st.download_button(
                label="Download Excel workbook",
                data=output.getvalue(),
                file_name=original_file_name.replace('.ifc', '.xlsx'),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

           # write_button = st.button("Download Excel", key="download_excel", on_click=download_excel)
           
            #if st.download_button:
            #    st.success("Excel file creation completed!")
             #   st.warning("Check your download folder")

            st.header("Filter by class")
            classesfilter = session.DataFrame["Class"].value_counts().keys().to_list()
            class_selectorfilter = st.selectbox("Select Class", options=classesfilter or [], key="class_selector_filter")

            st.write(get_ifc_pandas_filter(class_selectorfilter))
            
                    
        
        with tab2:
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("#### Options")
                classes = session.DataFrame["Class"].value_counts().keys().tolist()
                class_selector = st.selectbox("Select Class", options=classes or [], key="class_selector")
                
                session["filtered_frame"] = pandashelper.filter_dataframe_per_class(session.DataFrame, session.class_selector)
                
                
                
                session["qtos"] = pandashelper.get_qsets_columns(session["filtered_frame"], "RevitQuantities", "Tekla Quantity", "BaseQuantitiy")

                
                if session["qtos"] is not None:
                    qto_selector = st.selectbox("Select Quantity Set", session.qtos, key='qto_selector')
                    quantities = pandashelper.get_quantities(session.filtered_frame, session.qto_selector)
                    st.selectbox("Select Quantity", quantities, key="quantity_selector")
                    st.radio('Split per', ['Level', 'Type'], key="split_options")
                else:
                    st.warning("No quantities ~ ask your file issuer to export some")

            with col2:
                st.markdown("#### Graph")
                if "quantity_selector" in session:
                    if session.quantity_selector == "Count":
                        total = pandashelper.get_total(session["filtered_frame"])
                        st.write(f"The total number of {session.class_selector} is {total}")
                    else:
                        st.subheader(f'{session.class_selector} {session.quantity_selector}')
                        graph = graph_maker.load_graph(
                            session.filtered_frame,
                            session.qto_selector,
                            session.quantity_selector,
                            session.split_options,
                        )
                        st.write(graph)


        with tab3:
            st.header("Create BIMTypeCodes")

            col1, col2 = st.columns(2)

            with col1:
                st.markdown("#### Specify BIMTypeCodes")
                bim_type_codes = [
                    "TypeName",
                    "TypeCode",
                    "TypeNumber",
                    "TypeID",
                    "TypeDescription",
                    "Typetekst"
                ]
                bim_type_codes_selected = []
                for bim_type_code in bim_type_codes:
                    checkbox_state = st.checkbox(bim_type_code, value=True, key=f"checkbox_{bim_type_code}")
                    if checkbox_state:
                        bim_type_codes_selected.append(bim_type_code)
                
                st.markdown("##### Selected BIMTypeCodes")
                st.write(pd.DataFrame({"BIMTypeCodes": bim_type_codes_selected}))



            with col2:
                st.markdown("#### Create BIMTypeCodes in file")
                execute_button = st.button("Execute property creation")

            #if execute_button:
             #   update_properties(bim_type_codes_selected)
              #  st.success("Property creation completed successfully!")
               # st.warning("Check your download folder for the updated file")
   
                if execute_button:
                    updated_file_bytes = update_properties(bim_type_codes_selected)
                    st.success("Property creation completed successfully!")
                    st.download_button(
                        label="Download updated IFC file",
                        data=updated_file_bytes,
                        file_name=updated_file_name,  # Provide the appropriate name
                        mime="application/octet-stream"
                    )
            


        # ...

        with tab4:
            st.header("Upload data to model")

            file = st.file_uploader(" ", type=['csv', 'xlsx', 'pickle'])

            if not file:
                st.write("Upload a .csv, .xlsx, or .pickle file to get started")
            else:
                if file.type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                    xls = load_workbook(file, read_only=False)

                    #xls = pd.ExcelFile(file)
                    sheet_names = xls.sheetnames

                    col1, col2 = st.columns(2)

                    with col1:
                        data_preview = pd.DataFrame()
                        for sheet_name in sheet_names:
                            df_sheet = pd.read_excel(file, sheet_name)
                            data_preview = pd.concat([data_preview, df_sheet])

                        st.markdown("Data from uploaded file")
                        st.write(data_preview)

                        identifier_column = "GlobalId"
                        group_column = "Class"
                        new_columns = compare_datasets(data_preview, session["DataFrame"], identifier_column)
                       
                        # Update specific_columns to include all columns that start with "Pset_PAA"
                        specific_columns = [
                            col for col in data_preview.columns if col.startswith("Pset_PAA")
                        ]


                        differing_values_dict = compare_specific_values(session["DataFrame"], data_preview, specific_columns, identifier_column)
                        new_properties_dict = {}  # Dictionary to store properties by GlobalId



                        if not list(new_columns) and not list(differing_values_dict):
                            st.warning("No new columns or differing values found in the uploaded file.")
                        else:
                            grouped_properties = data_preview.groupby([group_column, identifier_column])

                            for (class_value, element), properties in grouped_properties:
                                new_properties = properties[list(new_columns)].dropna(how="all", axis=1)
                                if not new_properties.empty:
                                    properties_dict = new_properties.to_dict("list")
                                    properties_dict[group_column] = class_value  # Add the group column to the dictionary
                                    new_properties_dict[element] = properties_dict

                            for element, properties in differing_values_dict.items():
                                if element in new_properties_dict:
                                    for key, value in properties.items():
                                        if key in new_properties_dict[element]:
                                            new_properties_dict[element][key].extend(value)
                                        else:
                                            new_properties_dict[element][key] = value
                                else:
                                    new_properties_dict[element] = properties

                            st.markdown("""---""")
                            for element, properties in new_properties_dict.items():
                                class_value = properties[group_column]
                                st.warning(f"IfcBuildingElement: {element} | Class: {class_value}")
                                for col, values in properties.items():
                                    if col != group_column:
                                        for value in values:
                                            st.info(f"{col}: {value}")

                        print("Complete dictionary of psets")
                        print(new_properties_dict)



                    
                    with col2:
                        if not list(new_columns):
                            st.info("No new properties found in the uploaded file.")
                        else:
                            button_text = "Add New Properties"

                            if st.button(button_text):
                                # Logic to handle the button click and add new properties
                                updated_file_bytes = add_new_properties(new_properties_dict)  # Call the function to add new properties

                                st.success("New properties added successfully.")
                            
                                st.download_button(
                                    label="Download updated IFC file",
                                    data=updated_file_bytes,
                                    file_name=updated_file_name,  # Provide the appropriate name
                                    mime="application/octet-stream"
                                )


    
execute()

